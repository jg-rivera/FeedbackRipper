import openpyxl
import re
import json
from pathlib import Path


class Former:
    def __init__(self):
        xlsx_file = Path("eval.xlsx")
        self.workbook = openpyxl.load_workbook(xlsx_file)
        self.sheet = self.workbook.active
        print("Max columns:", self.sheet.max_column, ", Max rows:", self.sheet.max_row)

    def read(self):
        name_columns = {}
        feedback_columns = {}
        for i, row in enumerate(self.sheet.iter_rows(values_only=True)):
            if i == 0:
                for column_index in range(len(row)):
                    column = row[column_index]
                    # If it is name column
                    if self._is_name_column(column):
                        if self._has_numbers(column):
                            trimmed_column = column[:-1]
                            name_columns[trimmed_column].append(column_index)
                            continue
                        # Append first index
                        name_columns[column] = [column_index]
            else:
                for name, indices in name_columns.items():
                    for column_index in indices:
                        column_value = row[column_index]

                        if name in feedback_columns:
                            if i - 1 == len(feedback_columns[name]):
                                feedback_columns[name].append([column_value])
                                continue

                            feedback_columns[name][i - 1].append(column_value)
                            continue

                        feedback_columns[name] = [[column_value]]

        #        print(json.dumps(feedback_columns))
        self._turn_into_paragraphs(feedback_columns)

    def _turn_into_paragraphs(self, feedback_columns):
        questions = [
            "During events, what important roles and responsibilities have this person accomplished?",
            "What are the things that you like about this person's work ethic?",
            "If any, what difficulties have you encountered when working with this person?",
            "During meetings, how does this person spark collaboration among teams?",
            "If there is an improvement that you could suggest on this person in terms of how he/she handles responsibilities in the org, what would it be?",
        ]

        for name, feedbacks in feedback_columns.items():

            answers = [[None] * 1 for i in range(len(questions))]

            for feedback in feedbacks:
                for question_index in range(len(feedback)):
                    answer = feedback[question_index]
                    first_slot = answers[question_index][0]
                    if first_slot is None:
                        answers[question_index][0] = answer
                        continue
                    answers[question_index].append(answer)

            print(f"\n{name}")

            for answer_index in range(len(answers)):
                print(questions[answer_index])

                for answer in answers[answer_index]:
                    print("*", answer)

    def _is_name_column(self, column):
        return "-" in column

    def _has_numbers(self, string):
        return bool(re.search(r"\d", string))


if __name__ == "__main__":
    former = Former()
    former.read()